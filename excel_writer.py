"""
Excel 보고서 자동 생성 모듈

result 시트의 CONCATENATE(schedule!...) 수식을 파싱해서
키워드 위치·이미지 슬롯·순위 셀을 동적으로 감지합니다.
템플릿 구조(키워드 수, 그리드 크기, 시작 행)가 달라도 자동 대응합니다.
"""

import glob
import math
import os
import re
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.utils import column_index_from_string, get_column_letter

try:
    from PIL import Image as PILImage
    PIL_OK = True
except ImportError:
    PIL_OK = False


def _merge_sheet_styles(original_xml, output_xml):
    """Preserve original worksheet style refs while keeping updated cell values."""
    orig_root = ET.fromstring(original_xml)
    out_root = ET.fromstring(output_xml)

    ns_uri = ""
    if out_root.tag.startswith("{"):
        ns_uri = out_root.tag.split("}")[0].strip("{")

    def q(name):
        return f"{{{ns_uri}}}{name}" if ns_uri else name

    def find(parent, name):
        return parent.find(q(name))

    def findall(parent, name):
        return parent.findall(q(name))

    orig_cols = find(orig_root, "cols")
    out_cols = find(out_root, "cols")
    if orig_cols is not None:
        if out_cols is not None:
            out_root.remove(out_cols)
        sheet_data = find(out_root, "sheetData")
        insert_at = list(out_root).index(sheet_data) if sheet_data is not None else 0
        out_root.insert(insert_at, orig_cols)

    orig_sheet_data = find(orig_root, "sheetData")
    out_sheet_data = find(out_root, "sheetData")
    if orig_sheet_data is None or out_sheet_data is None:
        return ET.tostring(out_root, encoding="utf-8", xml_declaration=True)

    orig_rows = {row.attrib.get("r"): row for row in findall(orig_sheet_data, "row")}
    for out_row in findall(out_sheet_data, "row"):
        row_ref = out_row.attrib.get("r")
        orig_row = orig_rows.get(row_ref)
        if orig_row is None:
            continue

        for attr in ("s", "customFormat", "ht", "customHeight"):
            if attr in orig_row.attrib:
                out_row.attrib[attr] = orig_row.attrib[attr]
            else:
                out_row.attrib.pop(attr, None)

        orig_cells = {cell.attrib.get("r"): cell for cell in findall(orig_row, "c")}
        for out_cell in findall(out_row, "c"):
            cell_ref = out_cell.attrib.get("r")
            orig_cell = orig_cells.get(cell_ref)
            if orig_cell is None:
                continue
            if "s" in orig_cell.attrib:
                out_cell.attrib["s"] = orig_cell.attrib["s"]
            else:
                out_cell.attrib.pop("s", None)

    return ET.tostring(out_root, encoding="utf-8", xml_declaration=True)


def _restore_template_styles(template_path, output_path):
    """Restore original workbook styles after openpyxl-safe loading/saving."""
    import zipfile
    import tempfile

    fd, tmp_zip = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)

    try:
        with zipfile.ZipFile(template_path, "r") as zorig, \
             zipfile.ZipFile(output_path, "r") as zout, \
             zipfile.ZipFile(tmp_zip, "w", zipfile.ZIP_DEFLATED) as znew:

            original_names = set(zorig.namelist())

            for item in zout.infolist():
                data = zout.read(item.filename)

                if item.filename == "xl/styles.xml" and item.filename in original_names:
                    data = zorig.read(item.filename)
                elif item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml") and item.filename in original_names:
                    data = _merge_sheet_styles(
                        zorig.read(item.filename),
                        data,
                    )

                znew.writestr(item, data)

        os.replace(tmp_zip, output_path)
    finally:
        if os.path.exists(tmp_zip):
            try:
                os.remove(tmp_zip)
            except OSError:
                pass


def _load_workbook_with_fallback(path, **kwargs):
    """Load normally first, then fall back to a compatibility repair path."""
    try:
        return openpyxl.load_workbook(path, **kwargs), False
    except (TypeError, IndexError):
        pass

    import zipfile
    import re as _re
    from io import BytesIO
    from openpyxl.styles.stylesheet import Stylesheet
    from openpyxl.xml.functions import fromstring
    import xml.etree.ElementTree as ET

    with zipfile.ZipFile(path, "r") as zin:
        fixed_styles = None
        max_valid_style = None

        try:
            styles_xml = zin.read("xl/styles.xml")
            root = ET.fromstring(styles_xml)
            ns = root.tag.split("}")[0].strip("{") if root.tag.startswith("{") else ""
            cellstyles = root.find(f"{{{ns}}}cellStyles") if ns else root.find("cellStyles")
            if cellstyles is not None:
                for child in list(cellstyles):
                    if "name" not in child.attrib:
                        child.set("name", "Normal")
            fixed_styles = ET.tostring(root, encoding="utf-8", xml_declaration=True)
            style_count = len(Stylesheet.from_tree(fromstring(fixed_styles)).cell_styles)
            max_valid_style = max(0, style_count - 1)
        except Exception:
            fixed_styles = None
            max_valid_style = None

        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    if fixed_styles is not None:
                        data = fixed_styles
                    else:
                        text = data.decode("utf-8", errors="ignore")
                        text = _re.sub(
                            r'<(?P<prefix>[A-Za-z0-9_]+:)?cellStyle\b(?=[^>]*?/>)((?!\bname=)[^>])*?/>',
                            lambda m: m.group(0)[:-2] + ' name="Normal"/>',
                            text,
                        )
                        data = text.encode("utf-8")
                elif max_valid_style is not None and item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                    text = data.decode("utf-8", errors="ignore")
                    text = _re.sub(
                        r'\bs="(\d+)"',
                        lambda m: f's="{min(int(m.group(1)), max_valid_style)}"',
                        text,
                    )
                    text = _re.sub(
                        r'\bstyle="(\d+)"',
                        lambda m: f'style="{min(int(m.group(1)), max_valid_style)}"',
                        text,
                    )
                    data = text.encode("utf-8")
                zout.writestr(item, data)

    buffer.seek(0)
    return openpyxl.load_workbook(buffer, **kwargs), True


def _load_workbook_safe(path, **kwargs):
    """Backward-compatible wrapper returning only the workbook."""
    wb, _used_fallback = _load_workbook_with_fallback(path, **kwargs)
    return wb


# 이미지 사방 여백 (픽셀). 값을 늘리면 여백이 넓어집니다.
_IMAGE_PAD_PX = 8
_EMU_PER_PX   = 9525   # 96 DPI 기준 1 pixel = 9525 EMU


# ── 유틸 ─────────────────────────────────────────────────────────────────

def best_rank(api, screen):
    """API 순위와 화면 순위 중 더 높은(숫자가 낮은) 값을 반환."""
    api_num    = int(str(api).strip())    if str(api).strip().isdigit()    else None
    screen_num = int(str(screen).strip()) if str(screen).strip().isdigit() else None
    if api_num is not None and screen_num is not None:
        return min(api_num, screen_num)
    if api_num is not None:
        return api_num
    if screen_num is not None:
        return screen_num
    return "미노출"


def _safe_filename(text):
    """main.py safe_filename() 과 동일한 변환."""
    value = str(text or "").strip()
    value = re.sub(r"\s+", "", value)
    value = re.sub(r'[\\:*?"<>|]+', "_", value)
    value = re.sub(r"_+", "_", value).strip(" ._")
    return value[:80].rstrip(" ._") or "키워드없음"


# ── 동적 레이아웃 감지 ────────────────────────────────────────────────────

def analyze_template_layout(wb):
    """
    result 시트의 CONCATENATE(schedule!...) 수식을 파싱해서
    각 슬롯의 위치(이미지 영역, 순위 셀)를 자동 감지합니다.

    반환값: list of slot dict (schedule 순서대로 정렬)
    {
        'keyword'      : str   – 키워드 원문
        'label_row'    : int   – 레이블 셀 행 (1-indexed)
        'label_col'    : int   – 레이블 셀 열 (1-indexed)
        'img_row_start': int   – 이미지 영역 시작 행 (1-indexed)
        'img_row_end'  : int   – 이미지 영역 끝 행 (1-indexed)
        'img_col_start': int   – 이미지 영역 시작 열 (1-indexed)
        'img_col_end'  : int   – 이미지 영역 끝 열 (1-indexed)
        'rank_row'     : int   – 순위 셀 행 (1-indexed)
        'rank_col'     : int|None – 순위 셀 열 (1-indexed, 없으면 None)
    }
    """
    ws_result   = wb["result"]
    ws_schedule = wb["schedule"]

    # ── 1. CONCATENATE 수식 셀 수집 ─────────────────────────────────────
    formula_cells = []
    for row in ws_result.iter_rows():
        for cell in row:
            if not cell.value:
                continue
            val = str(cell.value)
            # CONCATENATE(...) 방식과 &연산자 방식 모두 지원
            if "schedule!" not in val.lower():
                continue
            m = re.search(r"schedule!\$?([A-Z]+)\$?(\d+)", val, re.IGNORECASE)
            if not m:
                continue
            sched_col = column_index_from_string(m.group(1).upper())
            sched_row = int(m.group(2))
            keyword   = ws_schedule.cell(row=sched_row, column=sched_col).value
            if not keyword:
                continue
            formula_cells.append({
                "keyword"  : str(keyword).strip(),
                "label_row": cell.row,
                "label_col": cell.column,
                "sched_row": sched_row,
            })

    if not formula_cells:
        return []

    # ── 2. schedule 순서대로 정렬 ────────────────────────────────────────
    formula_cells.sort(key=lambda x: x["sched_row"])

    # ── 3. 레이블 행·열 집합 ─────────────────────────────────────────────
    label_rows = sorted(set(f["label_row"] for f in formula_cells))

    # 각 레이블 행에 속하는 열 목록 (정렬)
    cols_by_label_row: dict[int, list[int]] = {}
    for f in formula_cells:
        cols_by_label_row.setdefault(f["label_row"], []).append(f["label_col"])
    for lr in cols_by_label_row:
        cols_by_label_row[lr].sort()

    # ── 4. 이미지 행 범위 계산 ───────────────────────────────────────────
    # 블록 간격 = 연속 레이블 행 차이의 최솟값 (가장 짧은 블록 기준)
    if len(label_rows) >= 2:
        block_gaps = [label_rows[i+1] - label_rows[i] for i in range(len(label_rows)-1)]
        block_gap  = min(block_gaps)   # 가장 짧은 블록 간격
    else:
        block_gap = 15  # 기본값

    # 헤더 끝 행: 첫 레이블 행 이전에 내용이 있는 마지막 행
    header_end_row = 0
    first_label_row = label_rows[0]
    for r in range(1, first_label_row):
        for cell in ws_result[r]:
            if cell.value is not None:
                header_end_row = max(header_end_row, r)

    def _img_row_range(label_row: int) -> tuple[int, int]:
        row_idx = label_rows.index(label_row)
        end_row = label_row - 1
        if row_idx == 0:
            # 첫 블록: 헤더 직후부터 시작, 단 다른 블록과 같은 높이를 유지
            start_by_gap    = label_row - block_gap + 1
            start_by_header = header_end_row + 1
            start_row = max(start_by_gap, start_by_header, 1)
        else:
            start_row = label_rows[row_idx - 1] + 1
        return start_row, end_row

    # ── 5. 이미지 열 범위 계산 ───────────────────────────────────────────
    def _img_col_range(label_row: int, label_col: int) -> tuple[int, int]:
        cols = cols_by_label_row[label_row]
        col_idx = cols.index(label_col)
        start_col = label_col
        if col_idx + 1 < len(cols):
            end_col = cols[col_idx + 1] - 1
        else:
            # 마지막 열: 앞 슬롯들의 평균 너비 사용
            if len(cols) >= 2:
                avg_span = sum(cols[i+1] - cols[i] for i in range(len(cols)-1)) // (len(cols)-1)
            else:
                avg_span = 7
            end_col = label_col + avg_span - 1
        return start_col, end_col

    # ── 6. 순위 셀 감지 ──────────────────────────────────────────────────
    def _has_fill(cell) -> bool:
        """셀에 흰색/검정/투명이 아닌 배경색이 있으면 True."""
        fill = cell.fill
        if not fill or fill.patternType in (None, "none"):
            return False
        fg = fill.fgColor
        if fg is None:
            return False
        if fg.type == "rgb":
            return fg.rgb not in ("00000000", "FFFFFFFF", "FF000000", "00FFFFFF")
        if fg.type == "theme":
            return True   # 테마 색상은 기본적으로 유색으로 간주
        if fg.type == "indexed":
            return fg.indexed not in (0, 1, 64)  # 64=투명, 0≈검정, 1≈흰색
        return False

    def _find_rank_cell(label_row: int, img_col_start: int, img_col_end: int,
                        label_col: int) -> int | None:
        search_end = img_col_end + 3   # 약간의 여유
        # 1순위: 기존 값이 있는 셀
        for c in range(img_col_start, search_end + 1):
            if c == label_col:
                continue
            cell = ws_result.cell(row=label_row, column=c)
            if cell.value is None:
                continue
            if "schedule!" in str(cell.value).lower():
                continue
            return c
        # 2순위: 값은 없어도 배경색이 있는 셀 (새 템플릿 대응)
        for c in range(img_col_start, search_end + 1):
            if c == label_col:
                continue
            cell = ws_result.cell(row=label_row, column=c)
            if cell.value and "schedule!" in str(cell.value).lower():
                continue
            if _has_fill(cell):
                return c
        return None

    # ── 7. 슬롯 조립 ─────────────────────────────────────────────────────
    slots = []
    for f in formula_cells:
        lr = f["label_row"]
        lc = f["label_col"]
        r_start, r_end = _img_row_range(lr)
        c_start, c_end = _img_col_range(lr, lc)
        rank_col       = _find_rank_cell(lr, c_start, c_end, lc)

        slots.append({
            "keyword"      : f["keyword"],
            "label_row"    : lr,
            "label_col"    : lc,
            "img_row_start": r_start,
            "img_row_end"  : r_end,
            "img_col_start": c_start,
            "img_col_end"  : c_end,
            "rank_row"     : lr,
            "rank_col"     : rank_col,
        })

    # ── 8. 순위 열 보정 (3단계) ──────────────────────────────────────────
    # ① 같은 레이블 열의 다른 블록에서 감지된 rank_col 재사용
    rank_col_by_label_col: dict[int, int] = {}
    for s in slots:
        if s["rank_col"] is not None:
            rank_col_by_label_col[s["label_col"]] = s["rank_col"]
    for s in slots:
        if s["rank_col"] is None and s["label_col"] in rank_col_by_label_col:
            s["rank_col"] = rank_col_by_label_col[s["label_col"]]

    # ② 전체가 새 템플릿이라 여전히 None인 경우: 슬롯 끝에서 2열 앞으로 추정
    #    (대부분의 템플릿에서 rank 셀은 이미지 영역 오른쪽 끝부분에 위치)
    for s in slots:
        if s["rank_col"] is None:
            s["rank_col"] = max(s["label_col"] + 1, s["img_col_end"] - 2)

    return slots


# ── 공개 API ─────────────────────────────────────────────────────────────

def read_schedule_keywords(wb):
    """
    analyze_template_layout() 로 얻은 키워드 목록을 반환.
    반환: [(keyword_str, slot_index), ...]
    """
    seq_col, kw_col, header_row = _find_schedule_seq_kw_cols(wb)
    if seq_col is not None:
        ws = wb["schedule"]
        rows = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            seq_val = ws.cell(row=row_idx, column=seq_col).value
            try:
                seq_num = int(seq_val)
            except (TypeError, ValueError):
                continue

            kw = str(ws.cell(row=row_idx, column=kw_col).value or "").strip()
            if kw:
                rows.append((kw, seq_num))

        if rows:
            return sorted(rows, key=lambda x: x[1])

    slots = analyze_template_layout(wb)
    return [(s["keyword"], i) for i, s in enumerate(slots)]


def _find_schedule_seq_kw_cols(wb):
    """schedule 시트에서 '순' / '키워드' 헤더 열 번호와 헤더 행 번호를 반환.

    팀원마다 시트 레이아웃이 달라도 셀 값으로 탐색하므로 행 위치에 의존하지 않음.
    반환: (seq_col, kw_col, header_row) or (None, None, None)
    """
    if "schedule" not in wb.sheetnames:
        return None, None, None
    ws = wb["schedule"]
    for row in ws.iter_rows():
        seq_col = kw_col = header_row = None
        for cell in row:
            if not hasattr(cell, "column") or not hasattr(cell, "row"):
                continue
            val = str(cell.value or "").strip()
            if val == "순":
                seq_col = cell.column
                header_row = cell.row
            elif val == "키워드":
                kw_col = cell.column
        if seq_col and kw_col and header_row:
            return seq_col, kw_col, header_row
    return None, None, None


def read_schedule_by_week(wb, week):
    """schedule 시트에서 해당 주차(week)의 키워드 목록을 반환.

    week=1 → 순 1~5, week=2 → 순 6~10, ...
    반환: [(keyword, seq_num), ...]  순번 오름차순
    """
    seq_col, kw_col, header_row = _find_schedule_seq_kw_cols(wb)
    if seq_col is None:
        return []
    ws = wb["schedule"]
    week_start = (week - 1) * 5 + 1
    week_end = week * 5
    results = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        seq_val = ws.cell(row=row_idx, column=seq_col).value
        try:
            seq_num = int(seq_val)
        except (TypeError, ValueError):
            continue
        if week_start <= seq_num <= week_end:
            kw = str(ws.cell(row=row_idx, column=kw_col).value or "").strip()
            if kw:
                results.append((kw, seq_num))
    return sorted(results, key=lambda x: x[1])


def get_schedule_max_week(wb):
    """schedule 시트의 최대 순번을 기준으로 전체 주차 수를 반환."""
    seq_col, _kw_col, header_row = _find_schedule_seq_kw_cols(wb)
    if seq_col is None:
        return 4
    ws = wb["schedule"]
    max_seq = 0
    for row_idx in range(header_row + 1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=seq_col).value
        try:
            max_seq = max(max_seq, int(val))
        except (TypeError, ValueError):
            continue
    return max(1, math.ceil(max_seq / 5))


def find_capture_file(keyword, captures_dir, preferred_rank=None):
    """
    captures_dir 하위 폴더에서 [safe_keyword]_*.png 파일 탐색.
    preferred_rank가 있으면 해당 순위 파일을 우선 반환.
    그 외에는 숫자 순위가 가장 높은(숫자가 낮은) 파일을 반환.
    """
    safe_kw = _safe_filename(keyword)
    matches = glob.glob(os.path.join(captures_dir, "**", f"{safe_kw}_*.png"), recursive=True)
    if not matches:
        return None

    preferred = str(preferred_rank or "").strip()
    if preferred.isdigit():
        preferred_name = f"{safe_kw}_{preferred}.png"
        for path in matches:
            if os.path.basename(path) == preferred_name:
                return path

    def _priority(path):
        suffix = os.path.splitext(os.path.basename(path))[0][len(safe_kw) + 1:]
        return (0, int(suffix)) if suffix.isdigit() else (1, path)

    return sorted(matches, key=_priority)[0]


def _slot_pixel_size(ws, col_start_1, col_end_1, row_start_1, row_end_1):
    """슬롯 범위의 픽셀 크기 계산 (1-indexed 입력)."""
    width_px = sum(
        int((ws.column_dimensions[get_column_letter(c)].width or 8.0) * 7 + 5)
        for c in range(col_start_1, col_end_1 + 1)
    )
    height_px = sum(
        int((ws.row_dimensions[r].height or 15.0) * 96 / 72)
        for r in range(row_start_1, row_end_1 + 1)
    )
    return max(width_px, 10), max(height_px, 10)


def insert_captures_to_report(template_path, captures_dir, output_path, ranks=None,
                               progress_callback=None):
    """
    Excel 템플릿의 result 시트에 캡처 이미지와 순위를 삽입하고 output_path 에 저장.

    매개변수:
        template_path     : 원본 Excel 보고서 템플릿 경로
        captures_dir      : 캡처 이미지 폴더 (하위 업체 폴더 포함)
        output_path       : 완성된 보고서 저장 경로
        ranks             : {keyword: rank_value} 딕셔너리 (int 또는 '미노출')
        progress_callback : 슬롯마다 호출되는 함수 (current, total, keyword) → None

    반환값: list of (keyword, status_str)
    """
    wb, used_fallback = _load_workbook_with_fallback(template_path)

    for sheet in ("schedule", "result"):
        if sheet not in wb.sheetnames:
            raise ValueError(f"'{sheet}' 시트를 찾을 수 없습니다.")

    ws_result = wb["result"]

    # 레이아웃 동적 감지
    slots = analyze_template_layout(wb)
    if not slots:
        raise ValueError(
            "result 시트에서 CONCATENATE(schedule!...) 수식을 찾을 수 없습니다.\n"
            "템플릿 형식을 확인해주세요."
        )

    # 기존 이미지 초기화
    ws_result._images.clear()

    report    = []
    tmp_files = []
    total     = len(slots)

    for idx, slot in enumerate(slots):
        keyword   = slot["keyword"]
        if progress_callback:
            progress_callback(idx + 1, total, keyword)
        r0_1      = slot["img_row_start"]   # 1-indexed
        r1_1      = slot["img_row_end"]
        c0_1      = slot["img_col_start"]
        c1_1      = slot["img_col_end"]
        rank_row  = slot["rank_row"]
        rank_col  = slot["rank_col"]

        # 이미지 슬롯이 유효하지 않으면 스킵
        if r1_1 < r0_1 or c1_1 < c0_1:
            report.append((keyword, "슬롯 계산 오류"))
            continue

        preferred_rank = ranks.get(keyword) if ranks and keyword in ranks else None
        capture_path = find_capture_file(keyword, captures_dir, preferred_rank=preferred_rank)
        if not capture_path:
            report.append((keyword, "캡처 없음"))
            # 캡처가 없어도 순위는 기입
            if ranks and keyword in ranks and rank_col:
                ws_result.cell(row=rank_row, column=rank_col).value = ranks[keyword]
            continue

        # PIL 리사이즈 (패딩을 뺀 크기로)
        tw, th    = _slot_pixel_size(ws_result, c0_1, c1_1, r0_1, r1_1)
        img_w     = max(tw - 2 * _IMAGE_PAD_PX, 10)
        img_h     = max(th - 2 * _IMAGE_PAD_PX, 10)
        use_path  = capture_path
        pil_error = None
        if PIL_OK:
            try:
                pil_img  = PILImage.open(capture_path).convert("RGB")
                pil_img  = pil_img.resize((img_w, img_h), PILImage.LANCZOS)
                tmp_path = capture_path + "_xl_tmp.png"
                pil_img.save(tmp_path)
                tmp_files.append(tmp_path)
                use_path = tmp_path
            except Exception as e:
                pil_error = str(e)   # 원본 이미지로 폴백, 오류는 report에 기록

        # TwoCellAnchor 로 이미지 삽입 (0-based 변환, 사방 패딩 적용)
        try:
            pad_emu       = _IMAGE_PAD_PX * _EMU_PER_PX
            last_col_w_px = int((ws_result.column_dimensions[get_column_letter(c1_1)].width or 8.0) * 7 + 5)
            last_row_h_px = int((ws_result.row_dimensions[r1_1].height or 15.0) * 96 / 72)

            xl_img        = XLImage(use_path)
            anchor        = TwoCellAnchor()
            anchor.editAs = "twoCell"
            anchor._from  = AnchorMarker(col=c0_1 - 1, colOff=pad_emu,   row=r0_1 - 1, rowOff=pad_emu)
            anchor.to     = AnchorMarker(
                col    = c1_1 - 1,
                colOff = max(0, last_col_w_px * _EMU_PER_PX - pad_emu),
                row    = r1_1 - 1,
                rowOff = max(0, last_row_h_px * _EMU_PER_PX - pad_emu),
            )
            xl_img.anchor = anchor
            ws_result.add_image(xl_img)
            cell_ref = f"{get_column_letter(c0_1)}{r0_1}:{get_column_letter(c1_1)}{r1_1}"
            if pil_error:
                report.append((keyword, f"완료(리사이즈실패: {pil_error[:60]}) [{cell_ref}]"))
            else:
                report.append((keyword, f"완료 ({cell_ref})"))
        except Exception as exc:
            report.append((keyword, f"삽입 오류: {exc}"))

        # 순위 셀 기입
        if ranks and keyword in ranks and rank_col:
            ws_result.cell(row=rank_row, column=rank_col).value = ranks[keyword]

    tmp_out = output_path + ".tmp"
    try:
        wb.save(tmp_out)
        os.replace(tmp_out, output_path)
        if used_fallback:
            _restore_template_styles(template_path, output_path)
    except Exception:
        try:
            os.remove(tmp_out)
        except OSError:
            pass
        raise

    for tmp in tmp_files:
        try:
            os.remove(tmp)
        except Exception:
            pass

    return report
